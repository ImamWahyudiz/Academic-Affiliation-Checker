"""
Output handling module (Excel, CSV)
"""

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def save_to_excel_with_highlight(df: pd.DataFrame, output_path: str) -> None:
    """
    Save DataFrame to Excel with yellow highlighting for flagged rows.
    
    Args:
        df: DataFrame to save
        output_path: Output file path (.xlsx)
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Find Flag column index
    flag_col_idx = headers.index("Flag") + 1 if "Flag" in headers else None
    
    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_flagged = False
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Check if this row is flagged
            if col_idx == flag_col_idx and str(value).strip().lower() == "yes":
                is_flagged = True
        
        # Apply yellow highlight to flagged rows
        if is_flagged:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
    
    # Auto-adjust column widths (approximate)
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=min(100, len(df) + 1), min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    wb.save(output_path)


def save_results(df: pd.DataFrame, output_path: str, use_excel: bool = True) -> str:
    """
    Save results to file (Excel or CSV).
    
    Args:
        df: DataFrame to save
        output_path: Base output path
        use_excel: If True, save as Excel with highlighting
        
    Returns:
        Actual output path used
    """
    if use_excel and OPENPYXL_AVAILABLE:
        # Change extension to .xlsx
        if output_path.endswith('.csv'):
            output_path = output_path[:-4] + '.xlsx'
        elif not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
        
        save_to_excel_with_highlight(df, output_path)
    else:
        # Fallback to CSV
        if output_path.endswith('.xlsx'):
            output_path = output_path[:-5] + '.csv'
        df.to_csv(output_path, index=False)
    
    return output_path
