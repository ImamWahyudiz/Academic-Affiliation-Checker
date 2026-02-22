"""
Academic Affiliation Background Checker
========================================
Deep affiliation background check using OpenAlex API.
Checks both direct and indirect (co-author) affiliations 
with institutions from specified countries.

Author: Academic Data Mining Project
License: MIT
"""

import pandas as pd
import requests
import time
import argparse
import sys
from typing import List, Dict, Tuple, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============================================================================
# CONFIGURATION
# ============================================================================

# Default configuration
DEFAULT_CONFIG = {
    "input_file": "Data.csv",
    "output_file": "Vetted_Output.csv",
    "flagged_countries": ["IL", "IR"],  # ISO 3166-1 alpha-2 codes
    "max_works_to_check": 30,
    "api_delay": 0.2,
    "user_agent": "mailto:academic-checker@example.com"
}

# Country code to name mapping (common ones)
COUNTRY_NAMES = {
    "IL": "Israel",
    "IR": "Iran",
    "RU": "Russia",
    "CN": "China",
    "KP": "North Korea",
    "SY": "Syria",
    "CU": "Cuba",
    "BY": "Belarus",
    "VE": "Venezuela",
    "MM": "Myanmar",
    "AF": "Afghanistan",
    "IQ": "Iraq",
    "LY": "Libya",
    "SD": "Sudan",
    "YE": "Yemen",
    "US": "United States",
    "GB": "United Kingdom",
    "DE": "Germany",
    "FR": "France",
    "JP": "Japan",
    "KR": "South Korea",
    "AU": "Australia",
    "CA": "Canada",
    "IN": "India",
    "BR": "Brazil",
    "ID": "Indonesia",
    "MY": "Malaysia",
    "SG": "Singapore",
    "SA": "Saudi Arabia",
    "AE": "UAE",
    "QA": "Qatar",
    "EG": "Egypt",
    "TR": "Turkey",
    "PK": "Pakistan",
}

# ============================================================================
# GENERIC INSTITUTION BLACKLIST
# ============================================================================
# OpenAlex sometimes incorrectly tags generic institution names (like "Ministry
# of Education") with wrong country codes. These patterns are excluded from
# affiliation checks to prevent false positives.

GENERIC_INSTITUTION_PATTERNS = [
    "ministry of education",
    "ministry of science",
    "ministry of health",
    "ministry of",
    "department of education",
    "national science foundation",
    "government of",
    "state council",
]


def is_generic_institution(institution_name: str) -> bool:
    """
    Check if an institution name is too generic and should be excluded.
    
    Args:
        institution_name: Name of the institution
        
    Returns:
        True if institution should be excluded (generic), False otherwise
    """
    if not institution_name:
        return False
    
    name_lower = institution_name.lower().strip()
    
    for pattern in GENERIC_INSTITUTION_PATTERNS:
        if pattern in name_lower:
            return True
    
    return False


def verify_author_name(
    expected_first: str, 
    expected_last: str, 
    actual_name: str,
    threshold: float = 0.5
) -> bool:
    """
    Verify that the OpenAlex author name matches the expected name.
    
    This helps prevent false positives from ID mismatches where different
    authors with similar names get confused.
    
    Args:
        expected_first: Expected first name
        expected_last: Expected last name
        actual_name: Actual display name from OpenAlex
        threshold: Minimum match ratio (0.0-1.0)
        
    Returns:
        True if names match sufficiently, False otherwise
    """
    if not actual_name:
        return False
    
    import re
    
    def normalize(name: str) -> str:
        """Normalize name: lowercase, remove punctuation except hyphens."""
        return re.sub(r'[^\w\s\-]', '', name.lower().strip())
    
    def get_name_variants(name: str) -> set:
        """Get all reasonable variants of a name."""
        name = normalize(name)
        variants = {name}
        
        # Add version without hyphens (merged)
        if '-' in name:
            variants.add(name.replace('-', ''))
        
        # Add individual parts if hyphenated
        if '-' in name:
            for part in name.split('-'):
                if len(part) > 1:
                    variants.add(part)
        
        return variants
    
    expected_first = normalize(expected_first)
    expected_last = normalize(expected_last)
    actual_name_normalized = normalize(actual_name)
    
    # Split actual name into parts
    actual_parts = set()
    for part in actual_name_normalized.split():
        actual_parts.add(part)
        # Handle merged hyphenated names
        if '-' in part:
            actual_parts.add(part.replace('-', ''))
            for subpart in part.split('-'):
                if len(subpart) > 1:
                    actual_parts.add(subpart)
    
    # Also add the full name without spaces for cases like "Liangjiang" vs "Liang Jiang"
    actual_merged = actual_name_normalized.replace(' ', '').replace('-', '')
    
    # Get expected name variants
    expected_first_variants = get_name_variants(expected_first)
    expected_last_variants = get_name_variants(expected_last)
    
    # Check if last name matches (strict: must be a complete word/part)
    last_name_match = False
    for exp_last in expected_last_variants:
        if exp_last in actual_parts:
            last_name_match = True
            break
    
    # Check if first name matches
    first_name_match = False
    for exp_first in expected_first_variants:
        # Skip very short variants unless they're the only option
        if len(exp_first) < 2 and len(expected_first_variants) > 1:
            continue
            
        # Check full first name as complete word
        if exp_first in actual_parts:
            first_name_match = True
            break
            
        # Check if first name is part of a merged name in actual
        # e.g., "guang" in "guangcan"
        for actual_part in actual_parts:
            if len(exp_first) >= 3 and actual_part.startswith(exp_first):
                first_name_match = True
                break
        
        if first_name_match:
            break
    
    # Check initials if first name doesn't match yet
    if not first_name_match and len(expected_first) >= 1:
        first_initial = expected_first[0]
        for part in actual_parts:
            # Match single-letter initials or 2-char initials like "M."
            if len(part) <= 2 and part[0] == first_initial:
                first_name_match = True
                break
            # Also check if actual name starts with expected first initial
            if part[0] == first_initial and len(part) > len(expected_first) * 0.5:
                first_name_match = True
                break
    
    # Both must match for confirmation
    # Exception: if expected first name is very short (1-2 chars), prioritize last name match
    if len(expected_first.replace(' ', '')) <= 2 and last_name_match:
        return True
    
    return last_name_match and first_name_match


def verify_institution_match(
    expected_institution: str,
    author_data: Dict
) -> Tuple[bool, float]:
    """
    Verify that the author's institution history contains the expected institution.
    
    This helps distinguish between different people with the same name
    by checking if their institution matches the input data.
    
    Args:
        expected_institution: Institution name from input data
        author_data: Author profile from OpenAlex
        
    Returns:
        Tuple of (is_match, confidence_score)
        - is_match: True if institution found in history
        - confidence_score: 0.0-1.0 indicating match quality
    """
    if not expected_institution or not author_data:
        return True, 0.0  # No institution to verify, allow pass
    
    import re
    
    def normalize_inst(name: str) -> str:
        """Normalize institution name for comparison."""
        if not name:
            return ""
        # Lowercase, remove punctuation
        name = re.sub(r'[^\w\s]', '', name.lower().strip())
        # Remove common words that don't help matching
        remove_words = ['university', 'of', 'the', 'institute', 'college', 
                       'school', 'department', 'faculty', 'center', 'centre',
                       'national', 'state', 'technical', 'technology']
        words = name.split()
        filtered = [w for w in words if w not in remove_words and len(w) > 2]
        return ' '.join(filtered) if filtered else name
    
    def get_key_words(name: str) -> set:
        """Extract key identifying words from institution name."""
        normalized = normalize_inst(name)
        return set(normalized.split())
    
    expected_normalized = normalize_inst(expected_institution)
    expected_keywords = get_key_words(expected_institution)
    
    if not expected_keywords:
        return True, 0.0  # No meaningful keywords to check
    
    # Collect all institution names from author's history
    author_institutions = set()
    
    # From affiliations
    affiliations = author_data.get("affiliations") or []
    for aff in affiliations:
        inst = aff.get("institution") or {}
        inst_name = inst.get("display_name", "")
        if inst_name:
            author_institutions.add(inst_name)
    
    # From last_known_institutions
    last_insts = author_data.get("last_known_institutions") or []
    for inst in last_insts:
        if inst:
            inst_name = inst.get("display_name", "")
            if inst_name:
                author_institutions.add(inst_name)
    
    # Check for matches
    best_score = 0.0
    
    for inst_name in author_institutions:
        inst_normalized = normalize_inst(inst_name)
        inst_keywords = get_key_words(inst_name)
        
        # Check exact normalized match
        if expected_normalized == inst_normalized:
            return True, 1.0
        
        # Check if expected is contained in author's institution or vice versa
        if expected_normalized in inst_normalized or inst_normalized in expected_normalized:
            best_score = max(best_score, 0.9)
            continue
        
        # Check keyword overlap (Jaccard similarity)
        if expected_keywords and inst_keywords:
            intersection = expected_keywords & inst_keywords
            union = expected_keywords | inst_keywords
            if union:
                jaccard = len(intersection) / len(union)
                if jaccard > 0.3:  # At least 30% overlap
                    best_score = max(best_score, jaccard)
    
    return best_score >= 0.3, best_score


def verify_author_identity(
    expected_first: str,
    expected_last: str,
    expected_institution: str,
    author_data: Dict
) -> Tuple[bool, str]:
    """
    Comprehensive author identity verification using name AND institution.
    
    Args:
        expected_first: Expected first name
        expected_last: Expected last name
        expected_institution: Expected current institution
        author_data: Author profile from OpenAlex
        
    Returns:
        Tuple of (is_verified, reason)
    """
    if not author_data:
        return False, "No author data"
    
    actual_name = author_data.get("display_name", "")
    
    # Step 1: Check name match
    name_match = verify_author_name(expected_first, expected_last, actual_name)
    
    # Step 2: Check institution match
    inst_match, inst_score = verify_institution_match(expected_institution, author_data)
    
    # Decision logic:
    # 1. Name matches + Institution matches → VERIFIED
    # 2. Name matches + No institution data → VERIFIED (trust name)
    # 3. Name matches + Institution doesn't match → WARNING but allow (might be old data)
    # 4. Name doesn't match + Institution matches → REJECT (different person)
    # 5. Name doesn't match + Institution doesn't match → REJECT
    
    if name_match:
        if inst_match:
            return True, f"Name and institution verified (score: {inst_score:.2f})"
        elif not expected_institution:
            return True, "Name verified (no institution to check)"
        else:
            # Name matches but institution doesn't - allow with warning
            return True, f"Name verified, institution not found in history (may be outdated)"
    else:
        if inst_match and inst_score >= 0.7:
            # Different name but strong institution match - might still be wrong person
            return False, f"Name mismatch (expected: '{expected_first} {expected_last}', got: '{actual_name}')"
        else:
            return False, f"Name mismatch (expected: '{expected_first} {expected_last}', got: '{actual_name}')"


def interactive_country_selection() -> List[str]:
    """
    Display interactive menu for country selection.
    
    Returns:
        List of selected country codes
    """
    print("\n" + "=" * 70)
    print("SELECT COUNTRIES TO CHECK FOR AFFILIATIONS")
    print("=" * 70)
    
    # Display country codes in compact columns
    items = list(COUNTRY_NAMES.items())
    cols = 4  # Number of columns
    col_width = 17
    
    for i in range(0, len(items), cols):
        row = items[i:i+cols]
        row_str = "".join([f"{code}={name:<{col_width-3}}" for code, name in row])
        print(f"  {row_str}")
    
    print("-" * 70)
    print("Input: kode negara dipisah spasi, contoh: IL IR RU")
    print("-" * 70)
    
    while True:
        user_input = input(">>> ").strip().upper()
        
        if not user_input:
            print("[!] Masukkan minimal 1 kode negara.")
            continue
        
        codes = user_input.split()
        valid_codes = []
        invalid_codes = []
        
        for code in codes:
            if code in COUNTRY_NAMES:
                valid_codes.append(code)
            else:
                invalid_codes.append(code)
        
        if invalid_codes:
            print(f"[!] Kode tidak dikenal: {', '.join(invalid_codes)}")
        
        if valid_codes:
            return valid_codes
        else:
            print("[!] Tidak ada kode valid. Coba lagi.")


def save_to_excel_with_highlight(df: pd.DataFrame, output_path: str) -> None:
    """
    Save DataFrame to Excel with yellow highlighting for flagged rows.
    
    Args:
        df: DataFrame to save
        output_path: Output file path (.xlsx)
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Find Flag column index
    flag_col_idx = headers.index("Flag") + 1 if "Flag" in headers else None
    
    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_flagged = False
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Check if this row is flagged
            if col_idx == flag_col_idx and str(value).strip().lower() == "yes":
                is_flagged = True
        
        # Apply yellow highlight to flagged rows
        if is_flagged:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
    
    # Auto-adjust column widths (approximate)
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=min(100, len(df) + 1), min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    wb.save(output_path)


# OpenAlex API base URL
OPENALEX_BASE = "https://api.openalex.org"


# ============================================================================
# API FUNCTIONS
# ============================================================================

def get_headers(user_agent: str) -> Dict:
    """Get request headers with polite user agent."""
    return {"User-Agent": user_agent}


def search_openalex_author(first_name: str, last_name: str, headers: Dict) -> Optional[str]:
    """
    Search for an author in OpenAlex by name.
    
    Args:
        first_name: Author's first name
        last_name: Author's last name
        headers: Request headers
        
    Returns:
        OpenAlex ID (without URL prefix) or None if not found
    """
    query = f"{first_name}+{last_name}"
    url = f"{OPENALEX_BASE}/authors?search={query}"
    
    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        results = data.get("results", [])
        if results and len(results) > 0:
            author = results[0]
            openalex_id = author.get("id", "")
            if openalex_id:
                return openalex_id.replace("https://openalex.org/", "")
        return None
    except requests.RequestException as e:
        print(f"    [API ERROR] Search: {e}")
        return None


def get_author_profile(openalex_id: str, headers: Dict) -> Optional[Dict]:
    """
    Fetch author profile from OpenAlex.
    
    Args:
        openalex_id: OpenAlex author ID
        headers: Request headers
        
    Returns:
        Author data dict or None
    """
    if not openalex_id:
        return None
        
    url = f"{OPENALEX_BASE}/authors/{openalex_id}"
    
    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"    [API ERROR] Author profile: {e}")
        return None


def get_author_works(openalex_id: str, headers: Dict, per_page: int = 30) -> List[Dict]:
    """
    Fetch recent works from an author.
    
    Args:
        openalex_id: OpenAlex author ID
        headers: Request headers
        per_page: Number of works to fetch
        
    Returns:
        List of work dictionaries
    """
    if not openalex_id:
        return []
        
    url = f"{OPENALEX_BASE}/works"
    params = {
        "filter": f"author.id:{openalex_id}",
        "sort": "publication_date:desc",
        "per-page": per_page
    }
    
    try:
        response = requests.get(url, headers=headers, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        return data.get("results", [])
    except requests.RequestException as e:
        print(f"    [API ERROR] Author works: {e}")
        return []


# ============================================================================
# AFFILIATION CHECK FUNCTIONS
# ============================================================================

def get_country_name(country_code: str) -> str:
    """Get human-readable country name from ISO code."""
    if not country_code:
        return "Unknown"
    return COUNTRY_NAMES.get(country_code, country_code)


def check_direct_affiliation(
    author_data: Optional[Dict], 
    flagged_countries: List[str]
) -> Tuple[bool, List[str]]:
    """
    Check direct affiliations from author profile.
    
    Args:
        author_data: Author profile data from OpenAlex
        flagged_countries: List of country codes to flag
        
    Returns:
        Tuple of (is_flagged, list_of_evidence)
    """
    if not author_data:
        return False, []
    
    evidence = []
    
    # Check affiliations array (full history)
    affiliations = author_data.get("affiliations") or []
    for aff in affiliations:
        if not isinstance(aff, dict):
            continue
            
        institution = aff.get("institution") or {}
        if not isinstance(institution, dict):
            continue
            
        country_code = institution.get("country_code")
        inst_name = institution.get("display_name", "Unknown Institution")
        
        # Skip generic institutions that may have incorrect country codes
        if is_generic_institution(inst_name):
            continue
        
        if country_code and country_code in flagged_countries:
            years = aff.get("years") or []
            
            if years and isinstance(years, list) and len(years) > 0:
                year_str = f" ({min(years)}-{max(years)})"
            else:
                year_str = ""
                
            country_name = get_country_name(country_code)
            evidence.append(f"{inst_name} [{country_name}]{year_str}")
    
    # Check last_known_institutions (array)
    last_institutions = author_data.get("last_known_institutions") or []
    for inst in last_institutions:
        if not isinstance(inst, dict):
            continue
            
        inst_name = inst.get("display_name", "Unknown Institution")
        
        # Skip generic institutions
        if is_generic_institution(inst_name):
            continue
            
        country_code = inst.get("country_code")
        if country_code and country_code in flagged_countries:
            country_name = get_country_name(country_code)
            ev = f"{inst_name} [{country_name}] (Last Known)"
            if ev not in evidence:
                evidence.append(ev)
    
    # Fallback: check last_known_institution (singular/legacy field)
    last_inst = author_data.get("last_known_institution")
    if last_inst and isinstance(last_inst, dict):
        inst_name = last_inst.get("display_name", "Unknown Institution")
        
        # Skip generic institutions
        if not is_generic_institution(inst_name):
            country_code = last_inst.get("country_code")
            if country_code and country_code in flagged_countries:
                country_name = get_country_name(country_code)
                ev = f"{inst_name} [{country_name}] (Last Known)"
                if ev not in evidence:
                    evidence.append(ev)
    
    return len(evidence) > 0, evidence


def check_indirect_affiliation(
    works: List[Dict], 
    target_author_id: str,
    flagged_countries: List[str]
) -> Tuple[bool, List[str]]:
    """
    Check indirect affiliations through co-authors.
    
    Args:
        works: List of work data from OpenAlex
        target_author_id: The author's OpenAlex ID (to exclude from checking)
        flagged_countries: List of country codes to flag
        
    Returns:
        Tuple of (is_flagged, list_of_evidence)
    """
    if not works:
        return False, []
    
    evidence = []
    checked_institutions = set()
    
    for work in works:
        if not isinstance(work, dict):
            continue
            
        authorships = work.get("authorships") or []
        pub_year = work.get("publication_year", "")
        
        for authorship in authorships:
            if not isinstance(authorship, dict):
                continue
                
            # Get author info
            author = authorship.get("author") or {}
            if not isinstance(author, dict):
                continue
                
            author_id_raw = author.get("id")
            if not author_id_raw:
                continue
                
            author_id = str(author_id_raw).replace("https://openalex.org/", "")
            
            # Skip the target author themselves
            if author_id == target_author_id:
                continue
            
            # Check co-author's institutions
            institutions = authorship.get("institutions") or []
            for inst in institutions:
                if not isinstance(inst, dict):
                    continue
                    
                country_code = inst.get("country_code")
                inst_id = inst.get("id", "")
                inst_name = inst.get("display_name", "Unknown Institution")
                
                # Skip generic institutions that may have incorrect country codes
                if is_generic_institution(inst_name):
                    continue
                
                if country_code and country_code in flagged_countries:
                    if inst_id and inst_id not in checked_institutions:
                        checked_institutions.add(inst_id)
                        
                        coauthor_name = author.get("display_name", "Unknown Co-author")
                        country_name = get_country_name(country_code)
                        
                        ev = f"Co-author: {coauthor_name} at {inst_name} [{country_name}] ({pub_year})"
                        evidence.append(ev)
    
    return len(evidence) > 0, evidence


# ============================================================================
# MAIN PROCESSING
# ============================================================================

def process_candidate(
    index: int, 
    row: pd.Series, 
    df: pd.DataFrame, 
    total_rows: int,
    config: Dict,
    headers: Dict,
    display_num: int = 0
) -> bool:
    """
    Process a single candidate for background check.
    
    Args:
        index: DataFrame row index
        row: Row data
        df: DataFrame (modified in place)
        total_rows: Total number of rows
        config: Configuration dictionary
        headers: Request headers
        display_num: Display number for progress (optional)
        
    Returns:
        True if flagged, False otherwise
    """
    first_name = str(row.get("First Name", "")).strip()
    last_name = str(row.get("Last Name", "")).strip()
    openalex_id = str(row.get("OpenAlex_ID", "")).strip()
    current_institution = str(row.get("Current Institution", "")).strip()
    
    flagged_countries = config["flagged_countries"]
    max_works = config["max_works_to_check"]
    api_delay = config["api_delay"]
    
    progress = display_num if display_num > 0 else index + 1
    print(f"\n[{progress}/{total_rows}] Checking: {first_name} {last_name}")
    print(f"    OpenAlex ID: {openalex_id}")
    if current_institution:
        print(f"    Institution: {current_institution}")
    
    # Initialize results
    flag_result = "No"
    affiliation_type = "None"
    flag_evidence = ""
    
    # ========================================================================
    # STEP 0: Verify Author Identity (Name + Institution)
    # ========================================================================
    author_data = get_author_profile(openalex_id, headers)
    time.sleep(api_delay)
    
    if author_data:
        actual_name = author_data.get("display_name", "")
        
        # Use comprehensive identity verification
        is_verified, verify_reason = verify_author_identity(
            first_name, last_name, current_institution, author_data
        )
        
        if not is_verified:
            print(f"    [WARNING] {verify_reason}")
            print(f"    [SKIP] Skipping affiliation check due to potential ID mismatch")
            df.at[index, "Flag"] = "No"
            df.at[index, "Affiliation_Type"] = "None"
            df.at[index, "Flag_Evidence"] = f"ID Mismatch: OpenAlex shows '{actual_name}'"
            return False
        
        print(f"    OpenAlex Name: {actual_name} [{verify_reason}]")
    
    # ========================================================================
    # STEP 1: Direct Affiliation Check
    # ========================================================================
    print("    [STEP 1] Checking direct affiliations...")
    
    if author_data:
        is_direct, direct_evidence = check_direct_affiliation(author_data, flagged_countries)
        
        if is_direct:
            flag_result = "Yes"
            affiliation_type = "Direct"
            flag_evidence = "; ".join(direct_evidence)
            print(f"    [FLAG] DIRECT affiliation found!")
            for ev in direct_evidence[:5]:
                print(f"           - {ev}")
            
            df.at[index, "Flag"] = flag_result
            df.at[index, "Affiliation_Type"] = affiliation_type
            df.at[index, "Flag_Evidence"] = flag_evidence
            return True
        else:
            print("    [OK] No direct affiliation found")
    else:
        print("    [WARNING] Could not fetch author profile")
    
    # ========================================================================
    # STEP 2: Indirect Affiliation Check (Co-authors)
    # ========================================================================
    print(f"    [STEP 2] Checking indirect affiliations (last {max_works} works)...")
    
    works = get_author_works(openalex_id, headers, max_works)
    time.sleep(api_delay)
    
    if works:
        print(f"    Found {len(works)} works to check")
        
        is_indirect, indirect_evidence = check_indirect_affiliation(
            works, openalex_id, flagged_countries
        )
        
        if is_indirect:
            flag_result = "Yes"
            affiliation_type = "Indirect (Co-author)"
            flag_evidence = "; ".join(indirect_evidence[:5])
            print(f"    [FLAG] INDIRECT affiliation found through co-authors!")
            for ev in indirect_evidence[:3]:
                print(f"           - {ev}")
            if len(indirect_evidence) > 3:
                print(f"           - ... and {len(indirect_evidence) - 3} more")
        else:
            print("    [OK] No indirect affiliation found")
    else:
        print("    [INFO] No works found")
    
    # Update DataFrame
    df.at[index, "Flag"] = flag_result
    df.at[index, "Affiliation_Type"] = affiliation_type
    df.at[index, "Flag_Evidence"] = flag_evidence
    
    return flag_result == "Yes"


def ensure_openalex_ids(df: pd.DataFrame, config: Dict, headers: Dict) -> pd.DataFrame:
    """
    Ensure all rows have OpenAlex IDs, searching if necessary.
    
    Args:
        df: Input DataFrame
        config: Configuration dictionary
        headers: Request headers
        
    Returns:
        DataFrame with OpenAlex_ID column populated
    """
    if "OpenAlex_ID" not in df.columns:
        print("[INFO] Column 'OpenAlex_ID' not found. Searching for authors...")
        df["OpenAlex_ID"] = None
    
    api_delay = config["api_delay"]
    needs_search = df["OpenAlex_ID"].isna() | (df["OpenAlex_ID"] == "")
    
    if needs_search.sum() > 0:
        print(f"[INFO] Searching OpenAlex IDs for {needs_search.sum()} candidates...")
        
        for index, row in df[needs_search].iterrows():
            first_name = str(row.get("First Name", "")).strip()
            last_name = str(row.get("Last Name", "")).strip()
            
            if first_name and last_name:
                print(f"  Searching: {first_name} {last_name}...", end=" ")
                openalex_id = search_openalex_author(first_name, last_name, headers)
                
                if openalex_id:
                    df.at[index, "OpenAlex_ID"] = openalex_id
                    print(f"Found: {openalex_id}")
                else:
                    print("Not found")
                    
                time.sleep(api_delay)
        
        print("[INFO] OpenAlex ID search complete.")
    
    return df


def run_background_check(config: Dict) -> None:
    """
    Main function to run the background check.
    
    Args:
        config: Configuration dictionary
    """
    print("=" * 70)
    print("ACADEMIC AFFILIATION BACKGROUND CHECKER")
    print("=" * 70)
    
    # Display configuration
    country_display = ", ".join([
        f"{code} ({get_country_name(code)})" 
        for code in config["flagged_countries"]
    ])
    print(f"\n[CONFIG] Flagged countries: {country_display}")
    print(f"[CONFIG] Max works to check: {config['max_works_to_check']}")
    print(f"[CONFIG] API delay: {config['api_delay']}s")
    
    headers = get_headers(config["user_agent"])
    
    # Read input file
    input_file = config["input_file"]
    print(f"\n[INFO] Reading file: {input_file}")
    
    try:
        df = pd.read_csv(input_file)
        print(f"[INFO] Loaded {len(df)} rows.")
    except FileNotFoundError:
        print(f"[ERROR] File '{input_file}' not found!")
        return
    except Exception as e:
        print(f"[ERROR] Failed to read file: {e}")
        return
    
    # Ensure OpenAlex IDs exist
    total_rows = len(df)
    df = ensure_openalex_ids(df, config, headers)
    
    # Filter candidates with OpenAlex ID
    has_id = df["OpenAlex_ID"].notna() & (df["OpenAlex_ID"] != "")
    candidates_df = df[has_id]
    candidates_with_id = len(candidates_df)
    skipped_count = total_rows - candidates_with_id
    
    print(f"[INFO] Total rows: {total_rows}")
    print(f"[INFO] With OpenAlex ID: {candidates_with_id}")
    if skipped_count > 0:
        print(f"[INFO] Skipped (no ID): {skipped_count}")
    
    if candidates_with_id == 0:
        print("[INFO] No candidates to process.")
        return
    
    # Initialize result columns
    df["Flag"] = "No"
    df["Affiliation_Type"] = "None"
    df["Flag_Evidence"] = ""
    
    # Process each candidate
    flagged_count = 0
    direct_count = 0
    indirect_count = 0
    flagged_candidates = []
    
    print(f"\n[INFO] Starting background check for {candidates_with_id} candidates...")
    print("-" * 70)
    
    for i, (df_index, row) in enumerate(candidates_df.iterrows()):
        try:
            is_flagged = process_candidate(
                df_index, row, df, candidates_with_id, config, headers, i + 1
            )
            
            if is_flagged:
                flagged_count += 1
                aff_type = df.at[df_index, "Affiliation_Type"]
                
                if aff_type == "Direct":
                    direct_count += 1
                else:
                    indirect_count += 1
                
                flagged_candidates.append({
                    "Name": f"{row['First Name']} {row['Last Name']}",
                    "Type": aff_type,
                    "Evidence": str(df.at[df_index, "Flag_Evidence"])[:100]
                })
                
        except KeyboardInterrupt:
            print("\n\n[INTERRUPTED] Process stopped by user.")
            print("[INFO] Saving progress...")
            break
        except Exception as e:
            print(f"    [ERROR] Failed to process: {e}")
            continue
    
    # Save results
    output_file = config["output_file"]
    
    # Change extension to xlsx for Excel output
    if output_file.endswith('.csv'):
        xlsx_file = output_file.replace('.csv', '.xlsx')
    else:
        xlsx_file = output_file + '.xlsx'
    
    print("\n" + "-" * 70)
    print(f"[INFO] Saving results to: {xlsx_file}")
    
    try:
        # Save to Excel with highlighting
        save_to_excel_with_highlight(df, xlsx_file)
        print("[SUCCESS] File saved successfully!")
        print(f"[INFO] Flagged rows are highlighted in YELLOW")
    except Exception as e:
        print(f"[ERROR] Failed to save Excel: {e}")
        # Fallback to CSV
        print("[INFO] Falling back to CSV...")
        try:
            df.to_csv(output_file, index=False, encoding="utf-8-sig")
            print(f"[SUCCESS] CSV saved: {output_file}")
        except Exception as e2:
            print(f"[ERROR] Failed to save CSV: {e2}")
            return
    
    # Print summary
    print("\n" + "=" * 70)
    print("BACKGROUND CHECK SUMMARY")
    print("=" * 70)
    
    clean_count = candidates_with_id - flagged_count
    print(f"\nTotal rows loaded: {total_rows}")
    print(f"Candidates checked: {candidates_with_id}")
    print(f"Clean (no flags): {clean_count} ({clean_count/candidates_with_id*100:.1f}%)")
    print(f"Flagged: {flagged_count} ({flagged_count/candidates_with_id*100:.1f}%)")
    print(f"  - Direct affiliation: {direct_count}")
    print(f"  - Indirect (co-author): {indirect_count}")
    
    if flagged_candidates:
        print("\n" + "-" * 70)
        print("FLAGGED CANDIDATES:")
        print("-" * 70)
        
        for i, fc in enumerate(flagged_candidates, 1):
            print(f"\n{i}. {fc['Name']}")
            print(f"   Type: {fc['Type']}")
            print(f"   Evidence: {fc['Evidence']}...")
    
    print(f"\n[OUTPUT] {output_file}")
    print("\n[DONE] Background check complete!")
    print("=" * 70)


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Academic Affiliation Background Checker using OpenAlex API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python affiliation_checker.py -i candidates.csv -o results.csv -c IL IR
  python affiliation_checker.py -i data.csv -c RU CN KP --works 50
  python affiliation_checker.py -i hiring.csv -c IL IR RU --delay 0.5

Country codes use ISO 3166-1 alpha-2 format (e.g., IL=Israel, IR=Iran, RU=Russia)
        """
    )
    
    parser.add_argument(
        "-i", "--input",
        default=DEFAULT_CONFIG["input_file"],
        help=f"Input CSV file (default: {DEFAULT_CONFIG['input_file']})"
    )
    
    parser.add_argument(
        "-o", "--output",
        default=DEFAULT_CONFIG["output_file"],
        help=f"Output CSV file (default: {DEFAULT_CONFIG['output_file']})"
    )
    
    parser.add_argument(
        "-c", "--countries",
        nargs="*",
        default=None,
        help="Country codes to flag (if not provided, interactive menu will be shown)"
    )
    
    parser.add_argument(
        "--works",
        type=int,
        default=DEFAULT_CONFIG["max_works_to_check"],
        help=f"Max works to check for indirect affiliations (default: {DEFAULT_CONFIG['max_works_to_check']})"
    )
    
    parser.add_argument(
        "--delay",
        type=float,
        default=DEFAULT_CONFIG["api_delay"],
        help=f"Delay between API requests in seconds (default: {DEFAULT_CONFIG['api_delay']})"
    )
    
    parser.add_argument(
        "--email",
        default=DEFAULT_CONFIG["user_agent"],
        help="Email for API polite pool (recommended for faster rate limits)"
    )
    
    return parser.parse_args()


def main():
    """Main entry point."""
    args = parse_arguments()
    
    # Check if countries were provided via command line
    if args.countries is None or len(args.countries) == 0:
        # No countries specified, show interactive menu
        selected_countries = interactive_country_selection()
    else:
        selected_countries = [c.upper() for c in args.countries]
    
    # Confirm selection
    country_names = [COUNTRY_NAMES.get(c, c) for c in selected_countries]
    print(f"\n[CONFIRMED] Checking affiliations with: {', '.join(country_names)}")
    
    # Build configuration from arguments
    config = {
        "input_file": args.input,
        "output_file": args.output,
        "flagged_countries": selected_countries,
        "max_works_to_check": args.works,
        "api_delay": args.delay,
        "user_agent": args.email
    }
    
    run_background_check(config)


if __name__ == "__main__":
    main()
